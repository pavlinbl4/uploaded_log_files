from openpyxl import load_workbook
from openpyxl import Workbook
import os
from datetime import datetime
from pathlib import Path


def create_report_file():
    report_date = datetime.today().strftime("%Y-%m-%d")
    path_to_file = f'{report_folder}/{report_file_name}'

    if os.path.exists(path_to_file):
        wb = load_workbook(path_to_file)  # файл есть и открываю его
        ws = wb['images']
        # ws = wb.create_sheet(report_date)  # добавляю новую таблицу
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файа еще нет
        ws.title = 'images'  # если файда еще нет

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20  # задаю ширину колонки
    ws.column_dimensions['D'].width = 20

    ws['A1'] = 'image_name'  # create columns names
    ws['B1'] = 'Ъ'
    ws['C1'] = 'PXP'
    ws['D1'] = 'TASS'

    wb.save(path_to_file)
    return path_to_file


def write_to_file(path_to_file, image_name, column_name, line_number):
    wb = load_workbook(path_to_file)
    ws = wb["images"]
    ws[f'A{line_number}'] = image_name
    ws[f'{column_name}{line_number}'] = '&&&'
    wb.save(path_to_file)


def find_row_with_image_name(image_name, path_to_report_file):  # return row number containing image_name
    wb = load_workbook(path_to_report_file)
    ws = wb["images"]
    for cell in ws['A']:
        if cell.value == image_name:
            return cell.row
    return None


def find_last_row(path_to_report_file): # find last row with data
    wb = load_workbook(path_to_report_file)
    ws = wb["images"]
    return ws.max_row


if __name__ == '__main__':
    (Path.home() / "Documents" / "photo_upload_logs").mkdir(parents=True, exist_ok=True)
    report_folder = Path.home() / "Documents" / "photo_upload_logs"
    report_file_name = 'uploaded_images.xlsx'

    image_name = '20171024_PAV5943_test_22.JPG'
    column_name = 'B'

    path_to_report_file = create_report_file()

    row_number = find_row_with_image_name(image_name, path_to_report_file)

    if row_number is None: # if no image_name in "A" column write data to new column
        row_number = find_last_row(path_to_report_file) + 1

    write_to_file(path_to_report_file, image_name, column_name, row_number)
